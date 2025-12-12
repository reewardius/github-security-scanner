#!/usr/bin/env python3
import os
import csv
import subprocess
import requests
import argparse
import shutil
import re
from time import sleep
from datetime import datetime
import pandas as pd
import json
import xml.etree.ElementTree as ET
from xml.dom import minidom
import boto3
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# ======== Аргументы командной строки ========
parser = argparse.ArgumentParser(description="GitHub search + TruffleHog scan")
group = parser.add_mutually_exclusive_group(required=True)
group.add_argument("-k", "--keyword", help="Single keyword or comma-separated keywords to search (e.g., 'password' or 'password,api_key,secret')")
group.add_argument("-kf", "--keywords-file", help="File with keywords (one per line)")
parser.add_argument("-t", "--token", help="GitHub API token")
parser.add_argument("-o", "--output", default="github_secrets", help="Output filename without extension (default: github_secrets)")
parser.add_argument("--max-size", type=int, default=500, help="Max repo size in MB (default: 500MB)")
parser.add_argument("--min-year", type=int, default=2024, help="Minimum year of last commit (default: 2024, only scan repos updated in 2024+)")
parser.add_argument("--issues", action="store_true", help="Search only in GitHub issues instead of code")

# Форматы вывода (можно указать несколько)
parser.add_argument("--xlsx", action="store_true", help="Save results to Excel (XLSX) format")
parser.add_argument("--csv", action="store_true", help="Save results to CSV format")
parser.add_argument("--json", action="store_true", help="Save results to JSON format")
parser.add_argument("--xml", action="store_true", help="Save results to XML format")
parser.add_argument("--txt", action="store_true", help="Save results to TXT format (pipe-separated)")

# AWS SES параметры (опционально)
parser.add_argument("--email-sender", help="Sender email address (must be verified in SES)")
parser.add_argument("--email-recipient", help="Recipient email address")
parser.add_argument("--aws-region", default="eu-central-1", help="AWS region for SES (default: eu-central-1)")

args = parser.parse_args()

# ======== Настройки ========
GITHUB_TOKEN = args.token
OUTPUT_BASENAME = args.output
TEMP_DIR = "./temp_repos"
MAX_SIZE_KB = args.max_size * 1024  # KB
MIN_YEAR = args.min_year  # Минимальный год последнего коммита
CACHE_FILE = ".github_scanner_cache.json"  # Локальный кеш просканированных репозиториев

# Если не указан ни один формат, по умолчанию используем XLSX
if not any([args.xlsx, args.csv, args.json, args.xml, args.txt]):
    args.xlsx = True

# ======== Удаляем старую папку temp_repos ========
if os.path.exists(TEMP_DIR):
    shutil.rmtree(TEMP_DIR)
os.makedirs(TEMP_DIR, exist_ok=True)

HEADERS = {"Authorization": f"token {GITHUB_TOKEN}"} if GITHUB_TOKEN else {}

# ======== Список ключевых слов ========
if args.keyword:
    # Поддержка множественных ключевых слов через запятую
    # Пример: -k "password,api_key,secret"
    SEARCH_KEYWORDS = [kw.strip() for kw in args.keyword.split(',') if kw.strip()]
elif args.keywords_file:
    with open(args.keywords_file, "r", encoding="utf-8") as f:
        SEARCH_KEYWORDS = [line.strip() for line in f if line.strip()]

# ======== Вспомогательные функции ========
def print_info(message):
    print(f"[INFO] {message}")

def print_warn(message):
    print(f"[WARN] {message}")

def print_secret(repo_url, keyword="", detector_type="", file_path="", line_number="", raw_result="", issue_url="", issue_title="", issue_body_snippet=""):
    if issue_url:
        print(f"[ISSUE] Keyword: {keyword} | URL: {issue_url} | Title: {issue_title} | Snippet: {issue_body_snippet}")
    else:
        print(f"[SECRET] Keyword: {keyword} | Repo: {repo_url} | Detector: {detector_type} | File: {file_path}:{line_number} | Secret: {raw_result}")

# ======== GitHub Code Search (с пагинацией) ========
def search_github_code_all_pages(keyword):
    results = []
    page = 1
    per_page = 100
    while True:
        url = f"https://api.github.com/search/code?q={keyword}+in:file&per_page={per_page}&page={page}"
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            print_warn(f"GitHub API Error {response.status_code}: {response.text}")
            break
        data = response.json()
        items = data.get("items", [])
        if not items:
            break
        results.extend(items)
        if len(items) < per_page:
            break
        page += 1
        sleep(1)
    return results

# ======== GitHub Issues Search (с пагинацией) ========
def search_github_issues_all_pages(keyword):
    results = []
    page = 1
    per_page = 100
    while True:
        url = f"https://api.github.com/search/issues?q={keyword}+in:title,body+is:issue&per_page={per_page}&page={page}"
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            print_warn(f"GitHub API Error {response.status_code}: {response.text}")
            break
        data = response.json()
        items = data.get("items", [])
        if not items:
            break
        results.extend(items)
        if len(items) < per_page:
            break
        page += 1
        sleep(1)
    return results

def get_last_commit(repo_full_name):
    url = f"https://api.github.com/repos/{repo_full_name}/commits?per_page=1"
    response = requests.get(url, headers=HEADERS)
    if response.status_code != 200 or not response.json():
        return ("Unknown", "Unknown")
    commit_data = response.json()[0]
    
    # Пробуем получить автора, если не получается - берем коммитера
    # author - реальный автор коммита
    # committer - может быть "GitHub" для веб-интерфейса
    author_name = commit_data["commit"]["author"]["name"]
    
    # Если автор "GitHub" или пустой, пробуем получить login пользователя
    if author_name == "GitHub" or not author_name or author_name == "Unknown":
        if commit_data.get("author") and commit_data["author"].get("login"):
            author_name = commit_data["author"]["login"]
    
    date = commit_data["commit"]["author"]["date"]
    return (author_name, date)

def get_commit_year(date_string):
    """Извлекает год из даты коммита в формате ISO 8601"""
    try:
        # Формат: 2024-01-15T10:30:00Z
        year = int(date_string.split('-')[0])
        return year
    except:
        return None

def load_cache():
    """Загружает кеш просканированных репозиториев из JSON файла"""
    try:
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
                return {
                    'repos': set(cache_data.get('repos', [])),
                    'last_scan': cache_data.get('last_scan', None),
                    'scan_count': cache_data.get('scan_count', 0)
                }
    except Exception as e:
        print_warn(f"Warning: Could not load cache file: {e}")
    
    # Возвращаем пустой кеш если файл не существует или ошибка
    return {
        'repos': set(),
        'last_scan': None,
        'scan_count': 0
    }

def save_cache(repos_set, previous_cache):
    """Сохраняет кеш просканированных репозиториев в JSON файл"""
    try:
        cache_data = {
            'repos': list(repos_set),
            'last_scan': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'scan_count': previous_cache['scan_count'] + 1
        }
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, indent=2, ensure_ascii=False)
        print_info(f"Cache saved to {CACHE_FILE}")
    except Exception as e:
        print_warn(f"Warning: Could not save cache file: {e}")

def get_repo_size(repo_full_name):
    url = f"https://api.github.com/repos/{repo_full_name}"
    response = requests.get(url, headers=HEADERS)
    if response.status_code != 200:
        return None, None
    repo_data = response.json()
    return repo_data.get("size", None), repo_data.get("private", None)

# ======== Функции сохранения в разных форматах ========
def save_to_xlsx(data, filename):
    """Сохранение в Excel формат с красивым форматированием"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    df = pd.DataFrame(data)
    
    # Создаем Excel writer
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Results')
        
        # Получаем workbook и worksheet
        workbook = writer.book
        worksheet = writer.sheets['Results']
        
        # Стили для границ
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Форматирование заголовков
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Применяем стили к заголовкам
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Чередующиеся цвета строк и границы
        light_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            # Чередуем цвета
            fill = light_fill if row_idx % 2 == 0 else white_fill
            
            for cell in row:
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=False)
        
        # Цветовое выделение для колонки "Keyword" (если есть)
        keyword_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        keyword_font = Font(bold=True, color='000000')
        
        if 'Keyword' in df.columns:
            keyword_col_idx = df.columns.get_loc('Keyword') + 1
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                          min_col=keyword_col_idx, max_col=keyword_col_idx):
                for cell in row:
                    cell.fill = keyword_fill
                    cell.font = keyword_font
                    cell.border = thin_border
        
        # Цветовое выделение для колонки "Detector Type" (если есть)
        if 'Detector Type' in df.columns:
            detector_col_idx = df.columns.get_loc('Detector Type') + 1
            detector_colors = {
                'AWS': 'FFE699',      # Желтый
                'Private Key': 'F4B084',  # Оранжевый
                'Database': 'C6E0B4',     # Зеленый
                'Generic API': 'B4C7E7',  # Голубой
            }
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                cell = row[detector_col_idx - 1]
                detector_type = str(cell.value)
                
                # Определяем цвет по типу детектора
                color = 'FFFFFF'  # По умолчанию белый
                for key, val in detector_colors.items():
                    if key.lower() in detector_type.lower():
                        color = val
                        break
                
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell.border = thin_border
        
        # Цветовое выделение для колонки "Status" (если есть)
        if 'Status' in df.columns:
            status_col_idx = df.columns.get_loc('Status') + 1
            new_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Светло-зеленый
            new_font = Font(bold=True, color='006400')  # Темно-зеленый текст
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                          min_col=status_col_idx, max_col=status_col_idx):
                for cell in row:
                    if cell.value and '🆕 NEW' in str(cell.value):
                        cell.fill = new_fill
                        cell.font = new_font
                    cell.border = thin_border
        
        # Автоматическая ширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            
            # Ограничиваем максимальную ширину
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)
        
        # Закрепляем верхнюю строку (заголовки)
        worksheet.freeze_panes = 'A2'
        
        # Добавляем автофильтр
        worksheet.auto_filter.ref = worksheet.dimensions
    
    print_info(f"Results saved to {filename}")

def save_to_csv(data, filename):
    """Сохранение в CSV формат"""
    if not data:
        print_warn(f"No data to save to {filename}")
        return
    
    fieldnames = list(data[0].keys())
    with open(filename, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
    print_info(f"Results saved to {filename}")

def save_to_json(data, filename):
    """Сохранение в JSON формат"""
    with open(filename, "w", encoding="utf-8") as jsonfile:
        json.dump(data, jsonfile, indent=2, ensure_ascii=False)
    print_info(f"Results saved to {filename}")

def save_to_xml(data, filename):
    """Сохранение в XML формат"""
    root = ET.Element("results")
    
    for item in data:
        record = ET.SubElement(root, "record")
        for key, value in item.items():
            element = ET.SubElement(record, key.replace(" ", "_").lower())
            element.text = str(value) if value else ""
    
    # Форматирование XML для читабельности
    xml_str = ET.tostring(root, encoding='unicode')
    dom = minidom.parseString(xml_str)
    pretty_xml = dom.toprettyxml(indent="  ")
    
    with open(filename, "w", encoding="utf-8") as xmlfile:
        xmlfile.write(pretty_xml)
    print_info(f"Results saved to {filename}")

def save_to_txt(data, filename):
    """Сохранение в TXT формат (pipe-separated)"""
    if not data:
        print_warn(f"No data to save to {filename}")
        return
    
    with open(filename, "w", encoding="utf-8") as txtfile:
        # Заголовки
        fieldnames = list(data[0].keys())
        txtfile.write(" | ".join(fieldnames) + "\n")
        txtfile.write("-" * (len(" | ".join(fieldnames))) + "\n")
        
        # Данные
        for item in data:
            row = " | ".join(str(item.get(field, "")) for field in fieldnames)
            txtfile.write(row + "\n")
    
    print_info(f"Results saved to {filename}")

def save_results(data, basename, formats):
    """Сохранение результатов во всех указанных форматах"""
    if args.xlsx:
        save_to_xlsx(data, f"{basename}.xlsx")
    if args.csv:
        save_to_csv(data, f"{basename}.csv")
    if args.json:
        save_to_json(data, f"{basename}.json")
    if args.xml:
        save_to_xml(data, f"{basename}.xml")
    if args.txt:
        save_to_txt(data, f"{basename}.txt")

def send_email_report(subject, body_text, sender, recipient, aws_region, attachments_info=None):
    """Отправка email отчета через AWS SES"""
    try:
        # Создаем SES клиент
        ses_client = boto3.client("ses", region_name=aws_region)
        
        # Создаем email сообщение
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = recipient
        msg["Subject"] = subject
        
        # Добавляем текст письма
        msg.attach(MIMEText(body_text, "plain"))
        
        # Отправляем email
        response = ses_client.send_raw_email(
            Source=sender,
            Destinations=[recipient],
            RawMessage={"Data": msg.as_string()}
        )
        
        print_info(f"✅ Email sent successfully! MessageId: {response['MessageId']}")
        return True
        
    except Exception as e:
        print_warn(f"❌ Error sending email: {e}")
        return False

# ======== Основной пайплайн ========
start_time = datetime.now()

try:
    # Загружаем кеш из предыдущих сканирований
    previous_cache = load_cache()
    previous_repos = previous_cache['repos']
    
    if previous_repos:
        print_info(f"📦 Loaded cache: {len(previous_repos)} repos from previous scans")
        print_info(f"   Last scan: {previous_cache['last_scan']}")
        print_info(f"   Total scans: {previous_cache['scan_count']}")
    else:
        print_info("📦 No previous cache found - first scan")
    
    total_repos_count = 0
    matched_repos_count = 0
    skipped_repos_count = 0  # Счетчик пропущенных (уже просканированных в этом запуске)
    old_repos_count = 0  # Счетчик старых репозиториев (до MIN_YEAR)
    new_repos_count = 0  # Счетчик новых репозиториев (не было в предыдущих сканах)
    
    results_list = []
    public_list = []
    
    # Кеш уже просканированных репозиториев в текущем запуске (по URL)
    scanned_repos_cache = set()
    # Список новых репозиториев (не было в предыдущих сканах)
    new_repos_list = []
    
    for keyword in SEARCH_KEYWORDS:
        print_info(f"Processing keyword: '{keyword}'")
        
        if args.issues:
            issues_results = search_github_issues_all_pages(keyword)
            for issue in issues_results:
                issue_url = issue["html_url"]
                issue_title = issue["title"]
                snippet = issue.get("body", "")[:200].replace("\n", " ")
                results_list.append({
                    "Keyword": keyword,
                    "Github Link": "",
                    "Last Commiter": "",
                    "Date of Last Commit": "",
                    "Detector Type": "",
                    "File Path": "",
                    "Line Number": "",
                    "Raw Result": "",
                    "Issue URL": issue_url,
                    "Issue Title": issue_title,
                    "Issue Body Snippet": snippet
                })
                print_secret("", keyword=keyword, issue_url=issue_url, issue_title=issue_title, issue_body_snippet=snippet)
                matched_repos_count += 1
        else:
            code_results = search_github_code_all_pages(keyword)
            for item in code_results:
                repo_url = item["repository"]["html_url"]
                repo_full_name = item["repository"]["full_name"]
                total_repos_count += 1
                
                # Проверяем кеш текущего запуска - если репозиторий уже сканировался, пропускаем
                if repo_url in scanned_repos_cache:
                    skipped_repos_count += 1
                    print_info(f"⚡ Skipping already scanned repo: {repo_url}")
                    continue
                
                # Проверяем является ли репозиторий новым (не было в предыдущих сканах)
                is_new_repo = repo_url not in previous_repos
                if is_new_repo:
                    new_repos_count += 1
                    print_info(f"🆕 NEW repository detected: {repo_url}")
                
                # Добавляем в кеш текущего запуска
                scanned_repos_cache.add(repo_url)

                repo_size_kb, is_private = get_repo_size(repo_full_name)
                if is_private:
                    print_warn(f"Skipping repo (private): {repo_url}")
                    continue

                local_path = os.path.join(TEMP_DIR, repo_full_name.replace("/", "_"))
                if repo_size_kb and repo_size_kb > MAX_SIZE_KB:
                    size_mb = repo_size_kb / 1024
                    print_warn(f"Skipping repo (too large {size_mb:.2f} MB > {args.max_size} MB): {repo_url}")
                    public_list.append({
                        "Github Link": repo_url,
                        "Author": "",
                        "Date of Last Commit": "",
                        "Keyword": keyword,
                        "Status": "🆕 NEW" if is_new_repo else "Existing"
                    })
                    continue
                
                # Получаем информацию о последнем коммите ПЕРЕД клонированием
                commiter, last_commit_date = get_last_commit(repo_full_name)
                
                # Проверяем год последнего коммита
                commit_year = get_commit_year(last_commit_date)
                if commit_year and commit_year < MIN_YEAR:
                    old_repos_count += 1
                    print_warn(f"Skipping old repo (last commit: {commit_year}): {repo_url}")
                    public_list.append({
                        "Github Link": repo_url,
                        "Author": commiter,
                        "Date of Last Commit": last_commit_date,
                        "Keyword": keyword,
                        "Status": "🆕 NEW" if is_new_repo else "Existing"
                    })
                    continue

                try:
                    subprocess.run(["git", "clone", "--quiet", "--depth", "1", repo_url, local_path], check=True)
                except subprocess.CalledProcessError:
                    print_warn(f"Failed to clone {repo_url}")
                    continue

                public_list.append({
                    "Github Link": repo_url,
                    "Author": commiter,
                    "Date of Last Commit": last_commit_date,
                    "Keyword": keyword,
                    "Status": "🆕 NEW" if is_new_repo else "Existing"
                })
                
                # Сохраняем информацию о новом репозитории
                if is_new_repo:
                    new_repos_list.append({
                        "Github Link": repo_url,
                        "Author": commiter,
                        "Date of Last Commit": last_commit_date,
                        "Keyword": keyword
                    })

                try:
                    result = subprocess.run(
                        ["trufflehog", "filesystem", "--only-verified", local_path],
                        capture_output=True,
                        text=True,
                        check=False
                    )
                    output = result.stdout.splitlines()
                    current_detector = ""
                    current_file = ""
                    current_line = ""

                    for line in output:
                        line = line.strip()
                        dt_match = re.search(r"Detector Type:\s*(.*)", line)
                        rr_match = re.search(r"Raw result:\s*(.*)", line)
                        file_match = re.search(r"Path:\s*(.*)", line)
                        line_match = re.search(r"Line Number:\s*(\d+)", line)

                        if dt_match:
                            current_detector = dt_match.group(1).strip()
                        if file_match:
                            current_file = file_match.group(1).strip()
                        if line_match:
                            current_line = line_match.group(1).strip()
                        if rr_match:
                            raw_result = rr_match.group(1).strip()
                            if raw_result:
                                results_list.append({
                                    "Keyword": keyword,
                                    "Github Link": repo_url,
                                    "Last Commiter": commiter,
                                    "Date of Last Commit": last_commit_date,
                                    "Detector Type": current_detector,
                                    "File Path": current_file,
                                    "Line Number": current_line,
                                    "Raw Result": raw_result,
                                    "Issue URL": "",
                                    "Issue Title": "",
                                    "Issue Body Snippet": ""
                                })
                                print_secret(repo_url, keyword=keyword, detector_type=current_detector, 
                                           file_path=current_file, line_number=current_line, raw_result=raw_result)
                                matched_repos_count += 1

                    print_info(f"SUCCESS scanning repo: {repo_url}")

                except Exception as e:
                    print_warn(f"Error scanning repo {repo_url}: {e}")

                shutil.rmtree(local_path, ignore_errors=True)
                sleep(1)

    # ======== Сохраняем результаты ========
    print_info("Saving results...")
    save_results(results_list, OUTPUT_BASENAME, args)
    save_results(public_list, f"{OUTPUT_BASENAME}_public_repos", args)
    
    # Сохраняем отдельный файл с новыми репозиториями (если есть)
    if new_repos_list:
        print_info(f"Saving NEW repositories report ({len(new_repos_list)} repos)...")
        save_results(new_repos_list, f"{OUTPUT_BASENAME}_new_repos", args)
    
    # Обновляем кеш
    save_cache(scanned_repos_cache, previous_cache)

    # ======== Итоговая статистика ========
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    
    unique_repos_scanned = len(scanned_repos_cache)
    
    print_info(f"Total repositories found: {total_repos_count}")
    print_info(f"Unique repositories scanned: {unique_repos_scanned}")
    if new_repos_count > 0:
        print_info(f"🆕 NEW repositories (not in previous scans): {new_repos_count} repos")
    if old_repos_count > 0:
        print_info(f"⏰ Skipped (old, before {MIN_YEAR}): {old_repos_count} repos")
    if skipped_repos_count > 0:
        print_info(f"⚡ Skipped (already scanned): {skipped_repos_count} repos")
        saved_time = skipped_repos_count * 5  # Примерно 5 секунд на репозиторий
        print_info(f"⚡ Time saved by caching: ~{saved_time} seconds")
    if old_repos_count > 0:
        saved_time_old = old_repos_count * 5
        print_info(f"⏰ Time saved by year filter: ~{saved_time_old} seconds")
    print_info(f"Repositories with secrets found: {matched_repos_count}")
    print_info(f"Scan duration: {duration:.1f} seconds")
    print_info(f"Scan complete!")

    # ======== Отправка email отчета (если указаны параметры) ========
    if args.email_sender and args.email_recipient:
        print_info("Preparing email report...")
        
        # Определяем какие форматы были сохранены
        formats_saved = []
        if args.xlsx:
            formats_saved.append("XLSX")
        if args.csv:
            formats_saved.append("CSV")
        if args.json:
            formats_saved.append("JSON")
        if args.xml:
            formats_saved.append("XML")
        if args.txt:
            formats_saved.append("TXT")
        
        formats_str = ", ".join(formats_saved)
        
        # Формируем тему письма
        if matched_repos_count > 0:
            subject = f"🚨 GitHub Security Scan - {matched_repos_count} Secrets Found!"
        else:
            subject = f"✅ GitHub Security Scan - No Secrets Found"
        
        # Список файлов результатов
        files_list = f"""
Results Files (Formats: {formats_str}):
- {OUTPUT_BASENAME}.{formats_saved[0].lower()} (main results with secrets)
- {OUTPUT_BASENAME}_public_repos.{formats_saved[0].lower()} (all scanned repositories)
"""
        
        # Топ детекторов (если есть результаты)
        detector_stats = ""
        if results_list:
            detectors = {}
            for item in results_list:
                detector = item.get('Detector Type', 'Unknown')
                if detector:
                    detectors[detector] = detectors.get(detector, 0) + 1
            
            if detectors:
                detector_stats = "\n\nTop Detector Types:"
                for detector, count in sorted(detectors.items(), key=lambda x: x[1], reverse=True)[:5]:
                    detector_stats += f"\n  - {detector}: {count} findings"
        
        # Топ ключевых слов с находками
        keyword_stats = ""
        if results_list:
            keywords = {}
            for item in results_list:
                kw = item.get('Keyword', 'Unknown')
                if kw:
                    keywords[kw] = keywords.get(kw, 0) + 1
            
            if keywords:
                keyword_stats = "\n\nTop Keywords with Findings:"
                for kw, count in sorted(keywords.items(), key=lambda x: x[1], reverse=True)[:5]:
                    keyword_stats += f"\n  - {kw}: {count} secrets found"
        
        # Статистика производительности
        performance_stats = ""
        perf_lines = []
        
        if old_repos_count > 0:
            perf_lines.append(f"  ⏰ Old repositories skipped (before {MIN_YEAR}): {old_repos_count}")
            perf_lines.append(f"  ⏰ Time saved by year filter: ~{old_repos_count * 5} seconds")
        
        if skipped_repos_count > 0:
            perf_lines.append(f"  ⚡ Duplicate repositories skipped: {skipped_repos_count}")
            perf_lines.append(f"  ⚡ Time saved by caching: ~{skipped_repos_count * 5} seconds")
        
        if perf_lines:
            performance_stats = "\nPerformance Optimization:\n" + "\n".join(perf_lines) + "\n"
        
        body_text = f"""GitHub Security Scanner Report
{'='*60}

Scan Summary:
-------------
Total repositories found: {total_repos_count}
Unique repositories scanned: {unique_repos_scanned}
Repositories with secrets found: {matched_repos_count}
Total secrets detected: {len(results_list)}
Scan duration: {duration:.1f} seconds
{performance_stats}
{files_list}
{detector_stats}
{keyword_stats}

{'='*60}
Scan completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

⚠️ Please review the results files and take appropriate action for any secrets found.
Ensure all exposed credentials are rotated immediately!

This is an automated report from GitHub Security Scanner.
"""
        
        # Отправляем email
        send_email_report(
            subject=subject,
            body_text=body_text,
            sender=args.email_sender,
            recipient=args.email_recipient,
            aws_region=args.aws_region
        )

except KeyboardInterrupt:
    print_warn("\n⚠️ Scan interrupted by user")
    
    # Отправляем email об остановке (если настроен)
    if args.email_sender and args.email_recipient:
        try:
            send_email_report(
                subject="⚠️ GitHub Security Scan - Interrupted",
                body_text=f"""GitHub Security Scanner was interrupted by user.

Scan details:
- Interrupted at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- Repositories scanned before interruption: {total_repos_count}
- Secrets found before interruption: {matched_repos_count}

This is an automated notification from GitHub Security Scanner.
""",
                sender=args.email_sender,
                recipient=args.email_recipient,
                aws_region=args.aws_region
            )
        except:
            pass
    
    exit(1)

except Exception as e:
    error_msg = f"Critical error: {str(e)}"
    print_warn(f"\n❌ {error_msg}")
    
    # Отправляем email об ошибке (если настроен)
    if args.email_sender and args.email_recipient:
        try:
            send_email_report(
                subject="❌ GitHub Security Scan - Critical Error",
                body_text=f"""A critical error occurred during the GitHub security scan.

Error details:
{error_msg}

Scan details:
- Error occurred at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- Repositories scanned: {total_repos_count if 'total_repos_count' in locals() else 0}
- Secrets found: {matched_repos_count if 'matched_repos_count' in locals() else 0}

Please check the scanner logs for more details.

This is an automated error notification from GitHub Security Scanner.
""",
                sender=args.email_sender,
                recipient=args.email_recipient,
                aws_region=args.aws_region
            )
        except:
            pass
    
    exit(1)
